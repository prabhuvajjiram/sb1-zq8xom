from fastapi import UploadFile
import PyPDF2
from docx import Document
import openpyxl
from io import BytesIO
from ..core.logger import logger
from ..utils.text_chunker import TextChunker

async def parse_document(file: UploadFile) -> str:
    content = ""
    try:
        file_content = await file.read()
        file_type = file.content_type
        
        if file_type == "application/pdf":
            content = parse_pdf(file_content)
        elif file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            content = parse_docx(file_content)
        elif file_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            content = parse_xlsx(file_content)
        else:
            # For text files or unknown types, try to decode as text
            content = file_content.decode()
        
        return content
    except Exception as e:
        logger.error(f"Error parsing document: {e}")
        raise

def parse_pdf(content: bytes) -> str:
    pdf_file = BytesIO(content)
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    text = ""
    for page in pdf_reader.pages:
        text += page.extract_text() + "\n"
    return text

def parse_docx(content: bytes) -> str:
    docx_file = BytesIO(content)
    doc = Document(docx_file)
    return "\n".join([paragraph.text for paragraph in doc.paragraphs])

def parse_xlsx(content: bytes) -> str:
    xlsx_file = BytesIO(content)
    wb = openpyxl.load_workbook(xlsx_file)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        text += f"\nSheet: {sheet}\n"
        for row in ws.iter_rows(values_only=True):
            text += " | ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
    return text

async def parse_confluence_attachments(confluence, page_id: str, attachments: list, vector_store, chunker: TextChunker) -> int:
    processed_count = 0
    for attachment in attachments:
        try:
            # Download attachment
            content = confluence.get_attachment_by_id(attachment["id"])
            
            # Parse based on file type
            if attachment["title"].endswith(".pdf"):
                text = parse_pdf(content)
            elif attachment["title"].endswith(".docx"):
                text = parse_docx(content)
            elif attachment["title"].endswith(".xlsx"):
                text = parse_xlsx(content)
            else:
                continue
            
            # Store in vector database with chunks
            doc_id = f"confluence_attachment_{attachment['id']}"
            chunks = chunker.get_chunks_with_overlap(text)
            
            for i, chunk in enumerate(chunks):
                chunk_id = f"{doc_id}_chunk_{i}"
                await vector_store.add_document(
                    doc_id=chunk_id,
                    text=chunk["text"],
                    metadata={
                        "type": "confluence_attachment",
                        "filename": attachment["title"],
                        "page_id": page_id,
                        "chunk_info": chunk["metadata"]
                    }
                )
            
            processed_count += 1
        
        except Exception as e:
            logger.error(f"Error processing attachment {attachment['title']}: {e}")
            continue
            
    return processed_count